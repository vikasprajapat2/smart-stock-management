# backend/app/routes/product_excel.py

import io
import uuid
import random

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session
from io import BytesIO

from app.database import get_db
from app.models.category import Category
from app.models.inventory import Inventory
from app.models.product import Product
from app.models.warehouse import Warehouse
from app.utils.inventory_helpers import log_inventory_change
from app.utils.product_helpers import generate_barcode, generate_sku

router = APIRouter(
    prefix="/products",
    tags=["Product Excel"]
)

# ── Constants ────────────────────────────────────────────────────────────────

# Column order for the Products sheet
TEMPLATE_COLUMNS = [
    "product_name",
    "sku",
    "category_name",
    "selling_price",
    "reorder_level",
    "unit",
    "initial_stock",
]

# Fixed unit options — shown as a dropdown in the template
UNIT_OPTIONS = [
    "pcs",
    "kg",
    "gram",
    "ltr",
    "ml",
    "meter",
    "cm",
    "box",
    "pack",
    "set",
    "roll",
    "pair",
]

# Column widths (characters)
_COL_WIDTHS = {
    "product_name":  30,
    "sku":           18,
    "category_name": 24,
    "selling_price": 18,
    "reorder_level": 16,
    "unit":          12,
    "initial_stock": 18,
}

# Tooltip/comment text shown on each header cell
_COL_COMMENTS = {
    "product_name":  "Required — enter the full product name.",
    "sku":           "Optional — leave blank to auto-generate.",
    "category_name": "Select from dropdown (or type a new category).",
    "selling_price": "Numeric value — product selling price.",
    "reorder_level": "Numeric value — minimum stock before reorder alert.",
    "unit":          "Select from dropdown — unit of measurement.",
    "initial_stock": "Opening stock quantity to load into Main Warehouse.",
}

# Max data rows covered by validation dropdowns
_MAX_IMPORT_ROWS = 10_000


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_hidden_sheet(wb: Workbook, sheet_name: str, values: list[str]):
    """Create a hidden sheet and populate column A with the given values."""
    ws = wb.create_sheet(sheet_name)
    for row_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=1, value=value)
    ws.sheet_state = "hidden"
    return ws


def _add_dropdown(
    ws,
    col_letter: str,
    sheet_name: str,
    num_rows: int,
    error_title: str,
    error_msg: str,
):
    """
    Attach a list DataValidation dropdown to col_letter rows 2.._MAX_IMPORT_ROWS+1,
    sourced from 'sheet_name'!$A$1:$A$num_rows.
    Sheet name is always single-quoted so spaces/special chars are safe.
    """
    num_rows = max(num_rows, 1)  # guard against zero-row range
    formula = f"'{sheet_name}'!$A$1:$A${num_rows}"
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,       # False = arrow IS visible (openpyxl flag is inverted)
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_msg,
    )
    dv.sqref = f"{col_letter}2:{col_letter}{_MAX_IMPORT_ROWS + 1}"
    ws.add_data_validation(dv)


# ==============================
# DOWNLOAD TEMPLATE
# ==============================

@router.get("/template")
def download_product_template(db: Session = Depends(get_db)):
    """
    Return an Excel import template with:
    - Products sheet  : header + one sample row, frozen header, bold headers,
                        per-column tooltips, dropdowns on category_name & unit.
    - Categories sheet: hidden; populated live from the database so new
                        categories appear automatically on next download.
    - Units sheet     : hidden; fixed UNIT_OPTIONS list.
    """

    # ── Fetch live categories ────────────────────────────────────────────────
    db_categories = db.query(Category).order_by(Category.category_name).all()
    category_names = [c.category_name for c in db_categories]
    if not category_names:
        # Sensible fallback so the dropdown is never empty on a fresh install
        category_names = [
            "Electronics",
            "Raw Material",
            "Spare Part",
            "Finished Goods",
            "Consumable",
        ]

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # Hidden reference sheets (created before Products so Products stays active)
    _make_hidden_sheet(wb, "Categories", category_names)
    _make_hidden_sheet(wb, "Units", UNIT_OPTIONS)

    # ── Products sheet ───────────────────────────────────────────────────────
    prod_ws = wb.active
    prod_ws.title = "Products"

    # Styling primitives
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    data_font    = Font(name="Arial", size=10)
    data_align   = Alignment(vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ───────────────────────────────────────────────────────────
    prod_ws.row_dimensions[1].height = 24

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)

        # Cell value & style
        cell = prod_ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border

        # Column width
        prod_ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(col_name, 16)

        # Tooltip comment on the header cell
        comment_text = _COL_COMMENTS.get(col_name, "")
        if comment_text:
            comment = Comment(comment_text, "Smart Stock")
            comment.width  = 220
            comment.height = 60
            cell.comment   = comment

    # ── Sample data row ──────────────────────────────────────────────────────
    prod_ws.row_dimensions[2].height = 20
    sample = {
        "product_name":  "MacBook Pro",
        "sku":           "MBP-001",
        "category_name": category_names[0],
        "selling_price": 120000,
        "reorder_level": 5,
        "unit":          "pcs",
        "initial_stock": 100,
    }
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = prod_ws.cell(row=2, column=col_idx, value=sample[col_name])
        cell.font      = data_font
        cell.alignment = data_align
        cell.border    = cell_border

    # ── Dropdowns ────────────────────────────────────────────────────────────
    cat_col_letter  = get_column_letter(TEMPLATE_COLUMNS.index("category_name") + 1)
    unit_col_letter = get_column_letter(TEMPLATE_COLUMNS.index("unit") + 1)

    _add_dropdown(
        ws=prod_ws,
        col_letter=cat_col_letter,
        sheet_name="Categories",
        num_rows=len(category_names),
        error_title="Invalid Category",
        error_msg="Please select a category from the dropdown list.",
    )
    _add_dropdown(
        ws=prod_ws,
        col_letter=unit_col_letter,
        sheet_name="Units",
        num_rows=len(UNIT_OPTIONS),
        error_title="Invalid Unit",
        error_msg="Please select a unit from the dropdown list.",
    )

    # ── Freeze header row ─────────────────────────────────────────────────────
    prod_ws.freeze_panes = "A2"

    # ── Serialise ────────────────────────────────────────────────────────────
    output = BytesIO()
    try:
        wb.save(output)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate Excel template: {exc}",
        )
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_template.xlsx"},
    )


# ==============================
# EXPORT PRODUCTS
# ==============================

@router.get("/export")
def export_products(db: Session = Depends(get_db)):
    """Export all products to Excel, including category name."""

    products = db.query(Product).all()

    data = [
        {
            "id":             product.id,
            "product_name":   product.product_name,
            "sku":            product.sku,
            "barcode":        product.barcode,
            "category_name":  (
                product.category.category_name if product.category else ""
            ),
            "selling_price":  float(product.selling_price or 0),
            "reorder_level":  product.reorder_level,
            "unit":           product.unit,
            "is_active":      product.is_active,
        }
        for product in products
    ]

    df     = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Products")

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"},
    )


# ==============================
# IMPORT PRODUCTS
# ==============================

@router.post("/import")
def import_products(
    file: UploadFile = File(...),
    db:   Session    = Depends(get_db),
):
    """
    Bulk import products from an Excel file.

    Columns:
        product_name  — required
        sku           — optional; auto-generated when blank
        category_name — optional; category is created automatically if missing
        selling_price — defaults to 0
        reorder_level — defaults to 0
        unit          — defaults to 'pcs'
        initial_stock — defaults to 0; added to Main Warehouse inventory
    """

    # ── Read & parse ─────────────────────────────────────────────────────────
    contents = file.file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file or format: {exc}",
        )

    df = df.fillna({
        "product_name":  "Unknown Product",
        "sku":           "",
        "category_name": "",
        "selling_price": 0.0,
        "reorder_level": 0,
        "unit":          "pcs",
        "initial_stock": 0,
    })

    # ── Resolve Main Warehouse ────────────────────────────────────────────────
    main_warehouse = (
        db.query(Warehouse)
        .filter(Warehouse.warehouse_name == "Main Warehouse")
        .first()
    )
    if not main_warehouse:
        main_warehouse = Warehouse(
            warehouse_name="Main Warehouse", location="Main Location"
        )
        db.add(main_warehouse)
        db.commit()
        db.refresh(main_warehouse)

    # ── Process rows ─────────────────────────────────────────────────────────
    imported = 0

    try:
        for _, row in df.iterrows():
            prod_name     = str(row.get("product_name", "Unknown Product")).strip()
            input_sku     = str(row.get("sku", "")).strip()
            category_name = str(row.get("category_name", "")).strip()
            initial_stock = int(float(row.get("initial_stock", 0)))

            # ── Category (get or create) ──────────────────────────────────────
            category_id = None
            if category_name:
                category = (
                    db.query(Category)
                    .filter(Category.category_name == category_name)
                    .first()
                )
                if not category:
                    category = Category(category_name=category_name)
                    db.add(category)
                    db.flush()
                category_id = category.id

            # ── Product (get or create) ───────────────────────────────────────
            product = (
                db.query(Product)
                .filter(Product.product_name == prod_name)
                .first()
            )

            if not product:
                # SKU
                if input_sku:
                    if db.query(Product).filter(Product.sku == input_sku).first():
                        raise HTTPException(
                            status_code=400,
                            detail=f"SKU '{input_sku}' already exists. Fix the file and retry.",
                        )
                    sku = input_sku
                else:
                    sku = None
                    for _ in range(5):
                        candidate = generate_sku(prod_name, None)
                        if not db.query(Product).filter(Product.sku == candidate).first():
                            sku = candidate
                            break
                    if not sku:
                        sku = str(uuid.uuid4())[:10].upper()

                # Barcode
                barcode = None
                for _ in range(5):
                    candidate = generate_barcode()
                    if not db.query(Product).filter(Product.barcode == candidate).first():
                        barcode = candidate
                        break
                if not barcode:
                    barcode = str(random.randint(1_000_000_000, 9_999_999_999))

                product = Product(
                    product_name=prod_name,
                    category_id=category_id,
                    sku=sku,
                    barcode=barcode,
                    selling_price=float(row.get("selling_price", 0.0)),
                    reorder_level=int(float(row.get("reorder_level", 0))),
                    unit=str(row.get("unit", "pcs")),
                    is_active=True,
                )
                db.add(product)
                db.flush()

            # ── Inventory ─────────────────────────────────────────────────────
            if initial_stock > 0:
                inventory = (
                    db.query(Inventory)
                    .filter(
                        Inventory.product_id  == product.id,
                        Inventory.warehouse_id == main_warehouse.id,
                    )
                    .first()
                )

                if inventory:
                    old_qty            = inventory.quantity
                    inventory.quantity += initial_stock
                    new_qty            = inventory.quantity
                else:
                    old_qty   = 0
                    new_qty   = initial_stock
                    inventory = Inventory(
                        product_id=product.id,
                        warehouse_id=main_warehouse.id,
                        quantity=new_qty,
                    )
                    db.add(inventory)

                db.flush()

                log_inventory_change(
                    db=db,
                    product_id=product.id,
                    old_qty=old_qty,
                    new_qty=new_qty,
                    action="IMPORT",
                )

            imported += 1

        db.commit()

    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Database error during import: {exc}",
        )

    return {"message": f"{imported} products imported successfully"}